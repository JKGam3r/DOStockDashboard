# Should be used as "history" showing past results and performance
# NOT to determine if a stock is a buy or sell

from datetime import datetime

import pandas_datareader as web
import pandas as pd
import numpy as np

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, StockChart, Reference
from openpyxl.chart.axis import DateAxis, ChartLines
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.chart.data_source import NumData, NumVal

import xlsxwriter as xlw

# The first thing the user sees upon running the program
def starting_prompt():
    print("Welcome to the DO Stock Dashboard program, an application that will create a Microsoft Excel")
    print("dashboard based on your input.  Please follow the directions to continue.\n")

    create_excel_file()

# Gets the ticker symbol, start, and end dates for finding stock information
def get_info():
    # Name of the stock
    ticker_symbol = input("Ticker Symbol: ")

    print("For the start and end dates, format in dd/mm/yyyy.")
    print("\te.g. 15/12/2015 (15 December 2015)")
    print("\te.g. 23/6/2020 (23 June 2020)")

    # Get the start date
    str_start_date = input("Start Date: ")
    tuple_start_date = str_start_date.split('/')

    # Get the end date
    str_end_date = input("End Date: ")
    tuple_end_date = str_end_date.split('/')

    # Put into datetime objects
    start_date = datetime(int(tuple_start_date[2]), int(tuple_start_date[1]), int(tuple_start_date[0]))
    end_date = datetime(int(tuple_end_date[2]), int(tuple_end_date[1]), int(tuple_end_date[0]))

    return (ticker_symbol, start_date, end_date)

# Creates the excel file with the data
def create_excel_file():
    stock_info = get_info()

    # Go to yahoo and add data into an excel file
    excel_file_name = f"{stock_info[0]}_Dashboard.xlsx"
    temp_df = web.DataReader(stock_info[0], 'yahoo', stock_info[1], stock_info[2])
    df = temp_df.rename_axis(None, axis=1).reset_index()
    df.to_excel(excel_file_name)

    # Create the pivot tables from which the visuals will be created
    add_pivot_tables(df, excel_file_name)

# Adds in the pivot tables
def add_pivot_tables(df, excel_file_name):
    # A list storing information on each pivot table
    pivot_list = [

    ]

    # Open a new workbook with xlsxwriter
    wb = xlw.Workbook(excel_file_name, {'nan_inf_to_errors': True})

    # First add all of the data for each day
    all_data_table = df
    data_ws = wb.add_worksheet('Data')
    data_ws_rows = all_data_table.shape[0]
    data_ws_columns = all_data_table.shape[1]
    data_ws_cell_range = xlw.utility.xl_range(0, 0, data_ws_rows, data_ws_columns - 1)
    data_ws_header = [{'header': str(di)} for di in all_data_table.columns.tolist()]
    data_ws.add_table(data_ws_cell_range,
                 {'header_row': True, 'first_column': False, 'columns': data_ws_header, 'data': all_data_table.values.tolist()})

    # Create a volume pivot table
    volume_pivot_table = df.pivot_table(index=[df['Date'].dt.quarter], columns=[df['Date'].dt.year],
                                        values=['Volume'], aggfunc=['sum'])
    vol_name = 'VolumePivotTable'
    vol_ws = wb.add_worksheet(vol_name)
    vol_rows = volume_pivot_table.shape[0]
    vol_columns = volume_pivot_table.shape[1]
    vol_cell_range = xlw.utility.xl_range(0, 0, vol_rows, vol_columns - 1)
    vol_header = [{'header': str(di[2])} for di in volume_pivot_table.columns.tolist()]
    vol_ws.add_table(vol_cell_range,
                     {'header_row': True, 'first_column': False, 'columns': vol_header, 'data': volume_pivot_table.values.tolist()})
    pivot_list.append((vol_name, vol_rows + 1, vol_columns))

    # Create a high-low pivot table
    temp_hl_table = df.pivot_table(index=[df['Date'].dt.year],
                                          aggfunc={'Open': 'mean', 'High': 'max', 'Low': 'min', 'Close': 'mean'})
    high_low_pivot_table = temp_hl_table.rename_axis(None, axis=1).reset_index()
    high_low_name = 'HighLowPivotTable'
    high_low_ws = wb.add_worksheet(high_low_name)
    high_low_rows = high_low_pivot_table.shape[0]
    high_low_columns = high_low_pivot_table.shape[1]
    high_low_cell_range = xlw.utility.xl_range(0, 0, high_low_rows, high_low_columns - 1)
    high_low_header = [{'header': str(di) + ' Price'} for di in high_low_pivot_table.columns.tolist()]
    high_low_ws.add_table(high_low_cell_range,
                     {'header_row': True, 'first_column': False, 'columns': high_low_header,
                      'data': high_low_pivot_table.values.tolist()})
    pivot_list.append((high_low_name, high_low_rows + 1, high_low_columns))

    # Close the workbook
    wb.close()

    # Access the same workbook, this time with openpyxl, to add in the charts
    wb_obj = openpyxl.load_workbook(excel_file_name)

    # Make the dashboard
    create_dashboard(wb_obj, excel_file_name, pivot_list)

# The main function which creates the dashboard
def create_dashboard(wb_obj, excel_file_name, pivot_list):
    # Create a new worksheet
    dashboard = wb_obj.create_sheet('Sheet_A')
    dashboard.title = 'Dashboard'

    # Hide the active/ data sheet
    wb_obj.active.sheet_state = 'hidden'

    # Set the dashboard to be the first sheet the user sees upon opening the file
    wb_obj.active = wb_obj['Dashboard']

    # All of the visuals that will be added to the dashboard
    volume_chart = volume_graph(wb_obj, pivot_list[0])
    high_low_chart = high_low_graph(wb_obj, pivot_list[1])

    # Add the charts
    dashboard.add_chart(volume_chart, anchor='A1')
    dashboard.add_chart(high_low_chart, anchor='J1')

    # Hide all pivot worksheets
    for entry in pivot_list:
        wb_obj[entry[0]].sheet_state = 'hidden'

    # Save all changes
    wb_obj.save(filename=excel_file_name)

# Displays a graph that displays the volume of shares traded according to time/ date
def volume_graph(wb_obj, graph_data):
    # Get the data worksheet the chart will be based on
    data_worksheet = wb_obj[graph_data[0]]

    # Create the chart
    line_chart = LineChart()
    line_chart.title = "Volume Per Quarter In Each Year"
    line_chart.x_axis.title = "Quarter"
    line_chart.y_axis.title = "Volume"

    # Add values and dates
    values = Reference(data_worksheet, min_col=1, min_row=1, max_col=graph_data[2], max_row=graph_data[1])

    # Add all values/ data and the proper titles
    line_chart.add_data(values, titles_from_data=True)

    # Add in markers to pinpoint each quarter
    for s in line_chart.series:
        s.marker.symbol = "circle"

    return line_chart

# Displays a graph that displays the open, high, low, and close values of shares traded
def high_low_graph(wb_obj, graph_data):
    # Get the data worksheet the chart will be based on
    data_worksheet = wb_obj[graph_data[0]]

    # Create the chart
    stock_chart = StockChart()
    stock_chart.title = "Open-High-Low-Close"
    #stock_chart.x_axis.title = "Value 1"
    #stock_chart.y_axis.title = "Value 2"

    # Add values and dates
    values = Reference(data_worksheet, min_col=2, min_row=1, max_col=graph_data[2], max_row=graph_data[1])
    labels = Reference(data_worksheet, min_col=1, min_row=2, max_col=1, max_row=graph_data[1])

    # Add all values/ data and the proper titles
    stock_chart.add_data(values, titles_from_data=True)
    stock_chart.set_categories(labels)

    for s in stock_chart.series:
        s.graphicalProperties.line.noFill = True

    stock_chart.hiLowLines = ChartLines()
    stock_chart.upDownBars = UpDownBars()

    # Due to an Excel bug, high-low lines will only be shown with this dummy data
    pts = [NumVal(idx=i) for i in range(len(values) - 1)]
    cache = NumData(pt=pts)
    stock_chart.series[-1].val.numRef.numCache = cache

    return stock_chart

starting_prompt()
