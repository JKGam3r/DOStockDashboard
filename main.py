''' Should be used as "history" showing past results and performance.
NOT to determine if a stock is a buy or sell. '''

import warnings # Used for ignoring future warnings

import datetime # Used for formatting dates

import pandas_datareader as web # Used for getting data from yahoo
import openpyxl # Used for creating the dashboard
import xlsxwriter as xlw # Used for adding tables to the excel file
import pandas as pd # Used for converting balance and income statements to dataframes
import requests # Used to access alpha vantage
from openpyxl.styles import Font, PatternFill # Used for changing the font of a cell and filling in the background

from StockDashboardApplication.program_code.graphs import *
from StockDashboardApplication.program_code.latest_statistics import *

# Specifically for getting rid of the 'week depreciated' future warning in pandas
warnings.simplefilter(action='ignore', category=FutureWarning)

# Used to access alpha vantage data
API_KEY = '' # INPUT YOUR ALPHA VANTAGE API KEY HERE

# Tested approximate dimensions of one standard Excel cell
CELL_WIDTH = 1.694
CELL_HEIGHT = 0.51

# Colors
BG_COLOR = 'D3D3D3' # normal background color
FG_COLOR = '000000' # text color
INFO_BG_COLOR = 'FFFFFF' # color for chart and text background

# The very first thing the user sees upon running the program
def starting_prompt():
    print("Welcome to the DO Stock Dashboard program, an application that will create a Microsoft Excel")
    print("dashboard based on your input.  Please follow the directions to continue.\n")

# Gets the ticker symbol, start, and end dates for finding stock information
def get_input_info():
    keep_iterating = True
    while keep_iterating:
        print("You can reset your input at any time by typing 'Reset' (no quotes).")
        print("You can exit the program at any time by typing 'Quit' (no quotes).")

        # The name of the stock
        ticker_symbol = input("Ticker Symbol: ").upper()
        if ticker_symbol == 'RESET':
            print()
            continue # Skip the rest of this iteration
        elif ticker_symbol == 'QUIT':
            keep_iterating = False
            return None

        # Put into datetime objects
        # Current day may not be over/ stock data still coming in; use day before
        end_date = datetime.datetime.today() - datetime.timedelta(days=1)
        start_date = end_date - datetime.timedelta(days=365)

        # Ask the user for a path
        path = input("Please enter a path (simply hit ENTER to put new file in same directory): ")
        cap_path = path.upper()
        if cap_path == 'RESET':
            print()
            continue
        elif cap_path == 'QUIT':
            keep_iterating = False
            return None

        # Append file name to path
        file_name = f'{ticker_symbol}_Stock_Dashboard.xlsx'
        if path != '':
            if path[-1] == '/' or path[-1] == '\\':
                path += file_name
            else:
                path += '\\' + file_name
        else:
            path += file_name

        return (ticker_symbol.upper(), start_date, end_date, (path, file_name))

# Creates the excel file with the data
def create_excel_file():
    # Get stock information
    stock_info = get_input_info()

    # No input/ user wants to exit
    if stock_info is None:
        return False

    ticker_symbol = stock_info[0]
    start_date = stock_info[1]
    end_date = stock_info[2]
    path = stock_info[3]

    temp_df = None # Original dataframe, stock prices

    is_df = None  # income statements
    bs_df = None # balance sheets

    try:
        temp_df = web.DataReader(ticker_symbol, 'yahoo', start_date, end_date)

        # Get the URLs
        is_url = f'https://www.alphavantage.co/query?function=INCOME_STATEMENT&symbol={ticker_symbol}&apikey={API_KEY}'
        bs_url = f'https://www.alphavantage.co/query?function=BALANCE_SHEET&symbol={ticker_symbol}&apikey={API_KEY}'

        # Request access to page
        r1 = requests.get(is_url)
        r2 = requests.get(bs_url)

        # Put data in JSon format
        is_data = r1.json()
        bs_data = r2.json()

        # Convert to pandas dataframes
        q_is_df = pd.DataFrame(is_data['quarterlyReports'])
        q_bs_df = pd.DataFrame(bs_data['quarterlyReports'])

        a_is_df = pd.DataFrame(is_data['annualReports'])
        a_bs_df = pd.DataFrame(bs_data['annualReports'])
    except:
        print(f"Could not retrieve data.  Possible fixes:")
        print(f"\t*Make sure the ticker symbol '{ticker_symbol}' exists.")
        print(f"\t*Make sure the end date ({end_date}) chronologically comes after the start date ({start_date}).")
        print(f"\t*Make sure there is data between the start and end dates.")
        return True

    # Successfully obtained data
    tdf = temp_df.rename_axis(None, axis=1).reset_index() # Dataframe with date as part of the table
    df = tdf.reindex(columns=['Date', 'Open', 'High', 'Low', 'Close', 'Volume', 'Adj Close'])

    # Add the new pivot tables
    add_pivot_tables(ticker_symbol, (df, q_is_df, q_bs_df, a_is_df, a_bs_df), path)

    # Excel file finished, continue for another run
    return True

# Makes a pivot table given the necessary data
def make_pivot_table(wb, table, name, header):
    ws = wb.add_worksheet(name)
    rows = table.shape[0]
    columns = table.shape[1]
    cell_range = xlw.utility.xl_range(0, 0, rows, columns - 1)
    ws.add_table(cell_range,
                          {'header_row': True, 'first_column': False, 'columns': header,
                           'data': table.values.tolist()})
    return (name, rows + 1, columns)

# Adds in the necessary pivot tables from which the graphs will be constructed
def add_pivot_tables(ticker_symbol, all_dataframes, excel_path):
    # Put all of the dataframes with all of the data into separate variables
    df = all_dataframes[0]
    q_is_df = all_dataframes[1]
    q_bs_df = all_dataframes[2]
    a_is_df = all_dataframes[3]
    a_bs_df = all_dataframes[4]

    # A list storing information on each pivot table
    pivot_list = []

    # A list of all the tables
    tables = [df, q_is_df, q_bs_df, a_is_df, a_bs_df]

    # Data from the original dataframe, according to time interval
    df_day = df['Date'].dt.day
    df_week = df['Date'].dt.week
    df_month = df['Date'].dt.month
    df_quarter = df['Date'].dt.quarter
    df_year = df['Date'].dt.year

    # Either the user's given path or the file name
    usable_path = ""

    # Open a new workbook with xlsxwriter
    wb = None
    try: # Try user's path
        wb = xlw.Workbook(excel_path[0], {'nan_inf_to_errors': True})
        usable_path = excel_path[0]
    except: # Put in same directory
        print("Could not find given directory.  Saving in current directory...")
        wb = xlw.Workbook(excel_path[1], {'nan_inf_to_errors': True})
        usable_path = excel_path[1]
        print("Successfully saved to directory.")

    # Initialize the tables
    row_end = df.shape[0] + 1
    standard_stock_table = df.loc[(row_end - 30):row_end, ['Date', 'Close']]
    standard_stock_table = standard_stock_table.reset_index(drop=True)

    open_high_low_close_table = df.loc[(row_end - 9):row_end, ['Date', 'Open', 'High', 'Low', 'Close']]
    open_high_low_close_table = open_high_low_close_table.reset_index(drop=True)

    # Add the pivot tables
    general_data_titles = ('Stock Data', 'Quarterly Income Statement Data', 'Quarterly Balance Sheet Data',
                           'Annual Income Statement Data', 'Annual Balance Sheet Data')

    make_pivot_table(wb, df, general_data_titles[0], [{'header': str(di)} for di in df.columns.tolist()])
    make_pivot_table(wb, q_is_df, general_data_titles[1], [{'header': str(di)} for di in q_is_df.columns.tolist()])
    make_pivot_table(wb, q_bs_df, general_data_titles[2], [{'header': str(di)} for di in q_bs_df.columns.tolist()])
    make_pivot_table(wb, a_is_df, general_data_titles[3], [{'header': str(di)} for di in a_is_df.columns.tolist()])
    make_pivot_table(wb, a_bs_df, general_data_titles[4], [{'header': str(di)} for di in a_bs_df.columns.tolist()])
    pivot_list.append(make_pivot_table(wb, standard_stock_table, 'Closing Table',
                                       [{'header': str(di)} for di in standard_stock_table.columns.tolist()]))
    pivot_list.append(make_pivot_table(wb, open_high_low_close_table, 'Open-High-Low-Close Table',
                                       [{'header': str(di)} for di in open_high_low_close_table.columns.tolist()]))

    tables.append(standard_stock_table)
    tables.append(open_high_low_close_table)

    # Close the workbook
    wb.close()

    # Make the dashboard
    create_dashboard(ticker_symbol, usable_path, general_data_titles, pivot_list, tables)

# The function responsible for making the visual (dashboard)
def create_dashboard(ticker_symbol, path, general_data_titles, pivot_list, tables):
    df = tables[0]
    q_is_df = tables[1]
    q_bs_df = tables[2]
    a_is_df = tables[3]
    a_bs_df = tables[4]
    standard_stock_table = tables[5]
    open_high_low_close_table = tables[6]

    # Create a new workbook object from the excel file created
    wb_obj = openpyxl.load_workbook(path)

    # Create the dashboard worksheet
    dashboard_worksheet = wb_obj.create_sheet('Sheet_A')
    dashboard_worksheet.title = 'Dashboard'

    # Hide the general data sheets
    for entry in general_data_titles:
        wb_obj[entry].sheet_state = 'hidden'

    # Set the dashboard to be the first sheet the user sees upon opening the file
    wb_obj.active = wb_obj['Dashboard']

    # Hide the gridlines for a cleaner look
    dashboard_worksheet.sheet_view.showGridLines = False

    # Add in the background
    cols = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O')
    ranges = (
        (cols, (1, 53)), # Background
        (cols[1:10], (26, 32)), # In between the charts
        (cols[7:10], (35, 51)), # Beside the bottom chart
        (cols[11:14], (4, 32)) # On the right side
    )
    color_background(dashboard_worksheet, BG_COLOR, INFO_BG_COLOR, ranges)

    # Add in data
    sst_last_row = standard_stock_table.shape[0] - 1
    last_close = round(float(standard_stock_table.at[sst_last_row, 'Close']), 2)
    df_last_row = df.shape[0] - 1
    volume = "{:,}".format(int(df.at[df_last_row, 'Volume']))
    day_range = (round(float(df.at[df_last_row, 'Low']), 2), round(float(df.at[df_last_row, 'High']), 2))
    year_range = (round(float(df[['Low']].min()), 2), round(float(df[['High']].max()), 2))
    is_date = q_is_df.at[0, 'fiscalDateEnding']
    revenue = "{:,}".format(int(q_is_df.at[0, 'totalRevenue']) - int(q_is_df.at[0, 'costOfRevenue']))
    net_income = "{:,}".format(int(q_is_df.at[0, 'netIncome']))

    add_stat_overview(dashboard_worksheet, 26, 2, last_close, volume, day_range, year_range, is_date, revenue, net_income)
    add_ohlc_stats(dashboard_worksheet, 36, 8, open_high_low_close_table)
    add_calculations(dashboard_worksheet, 5, 12, q_bs_df.at[0, 'fiscalDateEnding'],
                     a_is_df.at[0, 'fiscalDateEnding'], FinancialHealth(q_bs_df, q_is_df), Growth(a_is_df))
    add_special_thanks(dashboard_worksheet, 40, 12)

    # All of the visuals that will be added to the dashboard
    standard_chart = standard_stock_graph(wb_obj, pivot_list[0], ticker_symbol, CELL_WIDTH * 9, CELL_HEIGHT * 21, INFO_BG_COLOR)
    open_high_low_close_chart = open_high_low_close_graph(wb_obj, pivot_list[1], CELL_WIDTH * 5, CELL_HEIGHT * 18, INFO_BG_COLOR)

    # Set font and text for the header
    title_font_style = Font(size="14", bold=True, name='Arial', color=FG_COLOR)
    dashboard_worksheet.cell(row=1, column=2, value=f'{ticker_symbol} Dashboard').font = title_font_style
    sub_title_font_style = Font(size="18", name='Arial', color=FG_COLOR)
    close_amount = round(df.at[df.shape[0] - 1, 'Close'], 2)
    dashboard_worksheet.cell(row=2, column=2, value=f'${close_amount}').font = sub_title_font_style
    note_font_style = Font(size="8", italic=True, name='Arial', color=FG_COLOR)
    dashboard_worksheet.cell(row=2, column=4, value=f'* Price from last close.').font = note_font_style

    # Add the visuals
    dashboard_worksheet.add_chart(standard_chart, anchor='B4')
    dashboard_worksheet.add_chart(open_high_low_close_chart, anchor='B34')

    # Hide all pivot table worksheets
    for entry in pivot_list:
        wb_obj[entry[0]].sheet_state = 'hidden'

    # Save all changes
    wb_obj.save(filename=path)

# Paints the background with specific colors
def color_background(sheet, general_background, text_background, ranges):
    bg_range = ranges[0]
    # Fill in the background
    for c in bg_range[0]:
        for r in range(bg_range[1][0], bg_range[1][1]):
            sheet[f'{c}{r}'].fill = PatternFill(fgColor=general_background, fill_type="solid")

    # Fill in the parts where text will be
    txt_ranges = ranges[1:]
    for my_range in txt_ranges:
        for c in my_range[0]:
            for r in range(my_range[1][0], my_range[1][1]):
                sheet[f'{c}{r}'].fill = PatternFill(fgColor=text_background, fill_type="solid")

# Add in the statistcs overview
def add_stat_overview(sheet, row_start, col_start, last_close, volume, day_range, year_range, is_date, revenue, dividends):
    # Header
    font_style = Font(size="14", bold=True, name='Arial', color=FG_COLOR)
    sheet.cell(row=row_start, column=col_start, value='Statistics Overview').font = font_style

    # Notice message
    font_style = Font(size="6", italic=True, name='Arial', color=FG_COLOR)
    sheet.cell(row=row_start, column=col_start+3, value='* NOTICE: Stats do not include today\'s values.').font = font_style

    font_style = Font(size="9", name='Arial', color=FG_COLOR)
    spacing = '     '

    # Previous Close
    sheet.cell(row=row_start+2, column=col_start+1, value=f'Previous Close:{spacing}${last_close}').font = font_style

    # Volume Traded
    sheet.cell(row=row_start+3, column=col_start+1, value=f'Volume Traded:{spacing}{volume} shares').font = font_style

    # Day's Range
    sheet.cell(row=row_start+4, column=col_start+1, value=f'Day\'s Range:{spacing}${day_range[0]} - ${day_range[1]}').font = font_style

    # 52-Week Range
    sheet.cell(row=row_start+2, column=col_start+5, value=f'52-Week Range:{spacing}${year_range[0]} - ${year_range[1]}').font = font_style

    # Revenue
    sheet.cell(row=row_start+3, column=col_start+5, value=f'Revenue ({is_date}):{spacing}${revenue}').font = font_style

    # Dividend
    sheet.cell(row=row_start+4, column=col_start+5, value=f'Net Income ({is_date}):{spacing}${dividends}').font = font_style

# Adds in data for the open-high-low-close chart
def add_ohlc_stats(sheet, row_start, col_start, df):
    # Header
    header_font_style = Font(size="14", bold=True, name='Arial', color=FG_COLOR)
    sheet.cell(row=row_start, column=col_start, value='Volatility').font = header_font_style

    # Sub-header
    sub_header_font_style = Font(size="8", italic=True, name='Arial', color=FG_COLOR)
    sheet.cell(row=row_start+1, column=col_start, value='Values are = High - Low').font = sub_header_font_style

    rows = df.shape[0]

    font_style = Font(size="9", name='Arial', color=FG_COLOR)

    # Get difference for each row in the table
    for i in range(rows):
        full_date = df.at[i, 'Date']
        date = f'{full_date.month}-{full_date.day}'
        high = float(df.at[i, 'High'])
        low = float(df.at[i, 'Low'])
        diff = round(high - low, 2)

        sheet.cell(row=row_start+i+4, column=col_start, value=f'{date}:').font = font_style
        sheet.cell(row=row_start+i+4, column=col_start+2, value=f'${diff}').font = font_style

# Adds in calculations
def add_calculations(sheet, row_start, col_start, ratio_date, growth_date, financial_health, growth):
    # Get calculations
    quick_ratio = financial_health.quick_ratio()
    current_ratio = financial_health.current_ratio()
    debt_to_equity = financial_health.debt_to_equity()
    if quick_ratio == None:
        quick_ratio = 'Could not calculate.'
    else:
        quick_ratio = round(quick_ratio, 2)
    if current_ratio == None:
        current_ratio = 'Could not calculate.'
    else:
        current_ratio = round(current_ratio, 2)
    if debt_to_equity == None:
        debt_to_equity = 'Could not calculate.'
    else:
        debt_to_equity = round(debt_to_equity, 2)

    revenue_growth = growth.revenue_growth()
    operating_income_growth = growth.operating_income_growth()
    net_income_growth = growth.net_income_growth()
    if revenue_growth == None:
        revenue_growth = 'Could not calculate.'
    else:
        revenue_growth = round(revenue_growth, 2)
    if operating_income_growth == None:
        operating_income_growth = 'Could not calculate.'
    else:
        operating_income_growth = round(operating_income_growth, 2)
    if net_income_growth == None:
        net_income_growth = 'Could not calculate.'
    else:
        net_income_growth = round(net_income_growth, 2)

    # Header 1
    header_font_style = Font(size="14", bold=True, name='Arial', color=FG_COLOR)
    sheet.cell(row=row_start, column=col_start, value='Financial Health').font = header_font_style

    sub_header_font_style = Font(size="8", italic=True, name='Arial', color=FG_COLOR)
    sheet.cell(row=row_start+1, column=col_start, value=f'As of {ratio_date}').font = sub_header_font_style

    font_style = Font(size="9", name='Arial', color=FG_COLOR)

    # Quick Ratio
    sheet.cell(row=row_start+3, column=col_start, value=f'Quick Ratio: {quick_ratio}').font = font_style

    # Current Ratio
    sheet.cell(row=row_start+4, column=col_start, value=f'Current Ratio: {current_ratio}').font = font_style

    # Debt-to-Equity
    sheet.cell(row=row_start+5, column=col_start, value=f'Debt-to-Equity: {debt_to_equity}').font = font_style

    # Header 2
    sheet.cell(row=row_start+8, column=col_start, value='Growth').font = header_font_style

    # Sub header 2
    sheet.cell(row=row_start + 9, column=col_start, value=f'As of {growth_date}').font = sub_header_font_style

    # Revenue growth
    sheet.cell(row=row_start+11, column=col_start, value=f'Revenue Growth: {revenue_growth}%').font = font_style

    # Operating Income Growth
    sheet.cell(row=row_start+12, column=col_start,
               value=f'Operating Income Growth: {operating_income_growth}%').font = font_style

    # Net Income Growth
    sheet.cell(row=row_start + 13, column=col_start,
               value=f'Net Income Growth: {net_income_growth}%').font = font_style

# Add special thanks message
def add_special_thanks(sheet, row_start, col_start):
    font_style = Font(size="9", name='Arial', color=FG_COLOR)

    # Header
    sheet.cell(row=row_start, column=col_start, value='Special Thanks To:').font = font_style

    # List to thank
    sheet.cell(row=row_start+2, column=col_start, value='*   Yahoo').font = font_style
    sheet.cell(row=row_start+3, column=col_start, value='*   Alpha Vantage').font = font_style

    # Reason
    sheet.cell(row=row_start+5, column=col_start, value='... for providing company data.').font = font_style

# Code to run in the beginning
starting_prompt()

# Keep iterating until boolean is False
keep_running = True
while keep_running:
    keep_running = create_excel_file()
    response = input("Type 'Yes' (no quotes) to run again, all other responses will exit the program: ")
    if response.upper() != 'YES':
        keep_running = False
